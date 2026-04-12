"""
Estilo Excel alinhado ao tema glass do app (dark premium).
Aplica header escuro, texto claro, linhas alternadas e bordas para exports .xlsx.
"""
from __future__ import annotations

from typing import TYPE_CHECKING, List, Optional

if TYPE_CHECKING:
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

# Cores do tema (equivalente ao glass_card / theme_manager)
EXCEL_HEADER_BG = "1E293B"      # dark blue-grey (card bg)
EXCEL_HEADER_FONT = "E2E8F0"    # light text
EXCEL_ROW_EVEN = "F8FAFC"       # very light grey (alternating)
EXCEL_ROW_ODD = "FFFFFF"
EXCEL_BORDER = "94A3B8"         # slate-400
EXCEL_ACCENT = "6366F1"         # primary (optional use)


def apply_glass_theme_to_sheet(
    ws: "Worksheet",
    *,
    header_fill: bool = True,
    alternating_rows: bool = True,
    auto_width: bool = True,
    number_format_columns: Optional[List[str]] = None,
) -> None:
    """
    Aplica o tema glass (premium dark) a uma planilha openpyxl.
    - Linha 1: fundo escuro (#1E293B), fonte clara, negrito.
    - Demais linhas: alternância sutil de fundo, bordas finas.
    - Largura de colunas automática por conteúdo.
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color=EXCEL_BORDER)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    if header_fill and ws.max_row >= 1:
        header_fill_obj = PatternFill(start_color=EXCEL_HEADER_BG, end_color=EXCEL_HEADER_BG, fill_type="solid")
        header_font = Font(color=EXCEL_HEADER_FONT, bold=True, size=11)
        for cell in ws[1]:
            cell.fill = header_fill_obj
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    if alternating_rows and ws.max_row >= 2:
        fill_even = PatternFill(start_color=EXCEL_ROW_EVEN, end_color=EXCEL_ROW_EVEN, fill_type="solid")
        fill_odd = PatternFill(start_color=EXCEL_ROW_ODD, end_color=EXCEL_ROW_ODD, fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            is_even = (row_idx % 2) == 0
            for cell in row:
                cell.fill = fill_even if is_even else fill_odd
                cell.border = border
                cell.alignment = Alignment(horizontal="left", vertical="center")

    if number_format_columns and ws.max_row >= 1:
        header_row = [c.value for c in ws[1]]
        for col_name in number_format_columns:
            if col_name in header_row:
                col_idx = header_row.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f"{col_letter}{row_idx}"]
                    if cell.number_format == "General" and isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"

    if auto_width:
        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(
                (min(len(str(c.value or "")), 50) for c in column_cells),
                default=10,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = max(max_len + 2, 10)


def style_excel_workbook(workbook: "Workbook") -> None:
    """Aplica o tema glass a todas as planilhas do workbook."""
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        apply_glass_theme_to_sheet(ws)
